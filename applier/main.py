#!/usr/bin/env python3
"""
applier.main — orchestrator for the daily job-application pipeline.

Flow per run:
  1. load fresh jobs from remote_jobs.xlsx (the daily scrape output)
  2. compare against applications/_state.json to find unprocessed jobs
  3. for each new job, fetch the JD, tailor the resume + cover letter via Claude,
     publish artifacts to applications/<slug>/, optionally submit, update state
  4. save state and exit; the GH Actions step then commits and pushes everything

Configuration (env vars):
  ANTHROPIC_API_KEY       required for resume tailoring
  APPLIER_MODEL           override Claude model (default claude-sonnet-4-6)
  APPLIER_AUTO_SUBMIT     set to "1" to actually submit; otherwise dry-run
  APPLIER_MAX_PER_RUN     cap how many new jobs we process in a single run
  APPLIER_FIRST_NAME ...  candidate contact info (see applicator._candidate_payload)
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from applier.applicator import submit_application
from applier.jd_fetcher import fetch_jd
from applier.publisher import publish_application
from applier.resume_tailor import load_base_resume, load_profile, tailor_resume
from applier.slug import make_slug
from applier.state import ApplicationRecord, ApplicationState

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("applier")

EXCEL_FILE = Path("remote_jobs.xlsx")
SHEET_NAME = "Remote Jobs"
DEFAULT_MAX = int(os.environ.get("APPLIER_MAX_PER_RUN", "15"))


def _read_jobs_from_excel(path: Path) -> list[dict]:
    if not path.exists():
        logger.error("excel file not found: %s", path)
        return []
    wb = openpyxl.load_workbook(path, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        logger.error("sheet %r missing in %s", SHEET_NAME, path)
        return []
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    out = []
    for row in rows[1:]:
        record = dict(zip(header, row))
        url = record.get("URL")
        if not url:
            continue
        out.append(
            {
                "title": record.get("Title", "") or "",
                "company": record.get("Company", "") or "",
                "url": str(url),
                "source": record.get("Source", "") or "",
                "date_posted": record.get("Date Posted", "") or "",
            }
        )
    return out


def _uid_for_url(url: str) -> str:
    import hashlib

    return hashlib.md5(url.strip().rstrip("/").lower().encode()).hexdigest()


def _process_one(
    job: dict,
    state: ApplicationState,
    base_resume: str,
    profile: str,
) -> ApplicationRecord:
    uid = _uid_for_url(job["url"])
    record = state.get(uid)
    if record is None:
        from datetime import date

        try:
            posted = (
                date.fromisoformat(job["date_posted"])
                if isinstance(job["date_posted"], str) and job["date_posted"]
                else None
            )
        except ValueError:
            posted = None
        slug = make_slug(job["company"], job["title"], job["url"], posted)
        record = ApplicationRecord(
            uid=uid,
            slug=slug,
            company=job["company"],
            title=job["title"],
            url=job["url"],
            source=job["source"],
        )
        state.upsert(record)

    if record.state in {"submitted", "responded", "rejected", "skipped"}:
        return record

    try:
        if record.state == "queued":
            logger.info("[%s] fetching JD ...", record.slug)
            jd = fetch_jd(record.url)
            record.transition("jd_fetched", note=f"len={len(jd.raw_text)}")
        else:
            jd = fetch_jd(record.url)

        if record.state == "jd_fetched":
            logger.info("[%s] tailoring resume ...", record.slug)
            artifacts = tailor_resume(base_resume, jd.to_markdown(), profile)
            publish_application(record, jd, artifacts)
            record.transition("tailored", note="artifacts published")

        result = submit_application(record, jd)
        if result.submitted:
            record.transition(
                "submitted",
                note=f"portal={result.portal} confirmation={result.confirmation_id}",
            )
        else:
            record.history.append(
                {
                    "at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                    "from": record.state,
                    "to": record.state,
                    "note": f"submit deferred: {result.note}",
                }
            )
    except Exception as exc:  # noqa: BLE001 - we want to record the error and move on
        logger.exception("[%s] failed: %s", record.slug, exc)
        record.error = str(exc)[:500]
        record.transition("error", note=str(exc)[:200])

    return record


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Daily auto-applier")
    parser.add_argument(
        "--max",
        type=int,
        default=DEFAULT_MAX,
        help="max number of new jobs to process per run (default %(default)s)",
    )
    parser.add_argument(
        "--retry-errors",
        action="store_true",
        help="re-process records currently in 'error' state",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=EXCEL_FILE,
        help="path to remote_jobs.xlsx",
    )
    args = parser.parse_args(argv)

    base_resume = load_base_resume()
    profile = load_profile()
    state = ApplicationState()
    jobs = _read_jobs_from_excel(args.excel)
    logger.info("loaded %d jobs from %s", len(jobs), args.excel)

    targets = []
    for job in jobs:
        uid = _uid_for_url(job["url"])
        record = state.get(uid)
        if record is None:
            targets.append(job)
        elif args.retry_errors and record.state == "error":
            record.error = ""
            record.transition("queued", note="manual retry")
            targets.append(job)
        elif record.state in {"queued", "jd_fetched"}:
            # finish a previously interrupted run
            targets.append(job)
        if len(targets) >= args.max:
            break

    if not targets:
        logger.info("nothing new to process")
        state.save()
        return 0

    logger.info("processing %d job(s) (cap %d)", len(targets), args.max)
    submitted = tailored = errored = 0
    for job in targets:
        record = _process_one(job, state, base_resume, profile)
        if record.state == "submitted":
            submitted += 1
        elif record.state == "tailored":
            tailored += 1
        elif record.state == "error":
            errored += 1

    state.save()

    print("\n" + "=" * 60)
    print("  applier  |  ", datetime.now(timezone.utc).date().isoformat())
    print("=" * 60)
    print(f"  Targets       : {len(targets)}")
    print(f"  Tailored      : {tailored}")
    print(f"  Submitted     : {submitted}")
    print(f"  Errored       : {errored}")
    print(f"  Total tracked : {len(state)}")
    print("=" * 60 + "\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
