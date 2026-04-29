from __future__ import annotations

import logging
import os
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from scrapers.base import Job

logger = logging.getLogger(__name__)

SHEET_NAME = "Remote Jobs"
EXCEL_PATH = Path("remote_jobs.xlsx")

COLUMNS = [
    ("Date Posted", 14),
    ("Title", 45),
    ("Company", 30),
    ("Location", 25),
    ("Job Type", 18),
    ("Category", 28),
    ("Tags", 40),
    ("Salary", 22),
    ("Source", 18),
    ("URL", 60),
    ("Date Added", 14),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="EBF3FB")


class ExcelManager:
    def __init__(self, path: Path | str = EXCEL_PATH):
        self.path = Path(path)
        self._wb: Optional[Workbook] = None
        self._ws = None
        self._existing_uids: set[str] = set()
        self._new_count = 0
        self._load()

    # ── Internal ──────────────────────────────────────────────────────────────

    def _load(self) -> None:
        if self.path.exists():
            self._wb = openpyxl.load_workbook(self.path)
            if SHEET_NAME in self._wb.sheetnames:
                self._ws = self._wb[SHEET_NAME]
                self._collect_existing_uids()
                return
        # Fresh workbook
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.title = SHEET_NAME
        self._write_header()

    def _collect_existing_uids(self) -> None:
        """Read the URL column from the existing sheet and store hashed UIDs."""
        url_col_idx = self._col_index("URL")
        for row in self._ws.iter_rows(min_row=2, values_only=True):
            url = row[url_col_idx - 1]
            if url:
                import hashlib
                uid = hashlib.md5(str(url).strip().rstrip("/").lower().encode()).hexdigest()
                self._existing_uids.add(uid)

    def _col_index(self, name: str) -> int:
        for i, (col_name, _) in enumerate(COLUMNS, start=1):
            if col_name == name:
                return i
        raise KeyError(name)

    def _write_header(self) -> None:
        for col_idx, (name, width) in enumerate(COLUMNS, start=1):
            cell = self._ws.cell(row=1, column=col_idx, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            self._ws.column_dimensions[get_column_letter(col_idx)].width = width
        self._ws.row_dimensions[1].height = 22
        self._ws.freeze_panes = "A2"

    def _next_row(self) -> int:
        return self._ws.max_row + 1

    def _row_fill(self, row: int) -> Optional[PatternFill]:
        # Alternate row shading (row 1 = header)
        return ALT_FILL if row % 2 == 0 else None

    def _write_row(self, row: int, job: Job) -> None:
        today_str = datetime.now(timezone.utc).date().isoformat()
        values = [
            job.date_posted.isoformat(),
            job.title,
            job.company,
            job.location,
            job.job_type,
            job.category,
            job.tags_str(),
            job.salary,
            job.source,
            job.url,
            today_str,
        ]
        fill = self._row_fill(row)
        for col_idx, value in enumerate(values, start=1):
            cell = self._ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

        # Make URL a hyperlink
        url_col = self._col_index("URL")
        url_cell = self._ws.cell(row=row, column=url_col)
        url_cell.hyperlink = job.url
        url_cell.font = Font(color="0563C1", underline="single")

    # ── Public API ────────────────────────────────────────────────────────────

    def add_jobs(self, jobs: list[Job]) -> int:
        """Add jobs not already present. Returns count of newly added jobs."""
        # Group new jobs: newest date first
        new_jobs = [j for j in jobs if j.uid not in self._existing_uids]
        new_jobs.sort(key=lambda j: j.date_posted, reverse=True)

        # Insert new rows directly after header (row 2), pushing old rows down
        if new_jobs:
            self._ws.insert_rows(2, amount=len(new_jobs))
            for offset, job in enumerate(new_jobs):
                row_idx = 2 + offset
                self._write_row(row_idx, job)
                self._existing_uids.add(job.uid)

        self._new_count = len(new_jobs)
        logger.info("Added %d new jobs (skipped %d duplicates)", self._new_count, len(jobs) - self._new_count)
        return self._new_count

    def save(self) -> None:
        self._wb.save(self.path)
        logger.info("Saved workbook → %s", self.path)

    @property
    def total_rows(self) -> int:
        return max(0, self._ws.max_row - 1)  # subtract header

    @property
    def new_count(self) -> int:
        return self._new_count
